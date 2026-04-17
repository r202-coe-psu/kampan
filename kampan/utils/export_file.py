import io
import datetime
from bson.objectid import ObjectId
from openpyxl import Workbook
import pandas as pd
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

from kampan import models
from kampan.models import mas
from kampan.models.export_file import ExportFile
from kampan.utils.date_utils import format_date_th

EXPORT_MAS_COLUMN_MAP = {
    "mas_code": "รหัสงบประมาณ",
    "description": "รายละเอียด",
    "amount": "จำนวนเงินที่ขอจัดตั้ง",
    "used_amount": "จำนวนเงินที่ใช้ไปแล้ว",
    "remaining_amount": "จำนวนเงินคงเหลือ",
}
EXPORT_MAS_COLUMN_KEYS = list(EXPORT_MAS_COLUMN_MAP.keys())

PURCHASED_COLUMN_MAP = {
    "doc_date": "วันที่เอกสาร",
    "requisition_code": "เลขที่ มอ",
    "item": "รายการ",
    "amount": "จำนวนเงินขอเบิก",
    "quotation_winner": "ผู้รันเงิน",
}
PURCHASED_COLUMN_KEYS = list(PURCHASED_COLUMN_MAP.keys())


TIMELINE_ITEMS_COLUMN_MAP = {
    "requisition_code": "เลขที่รายการขอซื้อ",
    "running_number": "ลำดับ",
    "requisition_item": "รายการ",
    "seller": "ผู้ขาย",
    "serial_number": "เลขซีเรียล",
    "requisition_item_code": "เลขที่ใบเบิก",
    "insurance_start_date": "วันเริ่มประกัน",
    "insurance_end_date": "วันสิ้นสุดประกัน",
    "insurance_duration": "ระยะเวลาประกัน",
    "location": "สถานที่ใช้งาน",
    "responder_user": "ผู้รับผิดชอบ",
}
TIMELINE_ITEMS_COLUMN_KEYS = list(TIMELINE_ITEMS_COLUMN_MAP.keys())

# สร้าง style component ต่าง ๆ
font = Font(bold=True)  # ตัวหนา
alignment = Alignment(horizontal="center", vertical="center")  # จัดกลาง
border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

blue_bg = PatternFill(fill_type="solid", start_color="E6F2FF", end_color="E6F2FF")
light_blue_bg = PatternFill(fill_type="solid", start_color="DDEBF7", end_color="DDEBF7")


def format_year_range(start_date, end_date):
    if not start_date or not end_date:
        return ""

    start_year = start_date.year + 543
    end_year = end_date.year + 543

    if start_year == end_year:
        return f"ปี {start_year}"
    else:
        return f"ปี {start_year} - {end_year}"


def style_openpyxl_header(ws, headers, header_text, merge_cells):
    header_text = header_text

    # 3. Assign the text to the top-left cell of our target area (A1)
    ws["A1"] = header_text

    # 4. Merge cells from A1 to E1
    ws.merge_cells(merge_cells)

    # 5. Apply formatting to the merged cell
    # Note: When cells are merged, you apply styles to the top-left cell (A1)
    ws["A1"].alignment = Alignment(
        horizontal="center",
        vertical="center",
        wrap_text=True,  # Crucial for making the \n line breaks work
    )

    # วนทุกคอลัมน์ของ Header
    for col_idx, header_name in enumerate(headers, start=1):
        cell = ws.cell(row=2, column=col_idx)
        cell.value = header_name
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        cell.fill = blue_bg

    ws.row_dimensions[2].height = 30


def get_last_data_in_object(obj_data):
    """ดึงค่าข้อมูลสุดท้ายที่ไม่ใช่ค่าว่างจาก list"""
    for item in reversed(obj_data):
        if item not in (None, []):
            return item
    return ""


# ส่งออก mas
def get_purchased_items_in_mas(ws, mas_id, start_row, start_date, end_date):
    """
    ดึงข้อมูล reservation ที่เสร็จสิ้นทั้งหมดที่เกี่ยวข้องกับ MAS นี้
    แล้ว map ข้อมูลลง worksheet โดยจัดกลุ่มตาม requisition_code

    Layout per requisition:
        Row A : วันที่ order_confirmed | B: requisition_code | C: item1_name | D: item1_amount | E: quotation_winner
        Row A : (merged)               | B: (merged)         | C: item2_name | D: item2_amount | E: (merged)
    """
    mas = models.MAS.objects(id=mas_id).first()
    reservations = models.Reservation.objects(
        mas=mas, reservation_status="finished"
    ).order_by("-reserved_date")

    # --- group reservations by requisition ---
    # each reservation → 1 requisition → many items
    # collect: { requisition_id: { requisition, actual_amount, order_confirmed_timestamp } }
    grouped = {}
    for res in reservations:
        req = res.requisition
        if not req:
            continue
        req_id = str(req.id)

        # find order_confirmed timestamp from requisition_timeline
        # filter by start_date and end_date if provided
        start_dt = datetime.datetime.combine(start_date, datetime.time.min)
        end_dt = datetime.datetime.combine(end_date, datetime.time.max)

        timeline = models.RequisitionTimeline.objects(
            requisition=req,
            progress__match={
                "progress_status": "order_confirmed",
                "timestamp__gte": start_dt,
                "timestamp__lte": end_dt,
            },
        ).first()

        order_confirmed_ts = None
        if timeline:
            for p in timeline.progress:
                if p.progress_status == "order_confirmed":
                    order_confirmed_ts = p.timestamp
                    break
        else:
            # Skip this requisition if no matching timeline entry in date range
            continue

        quotation_winner = timeline.quotation_winner if timeline else None

        if req_id not in grouped:
            grouped[req_id] = {
                "requisition": req,
                "actual_amount": float(res.actual_amount or 0),
                "order_confirmed_ts": order_confirmed_ts,
                "quotation_winner": quotation_winner,
            }
        else:
            grouped[req_id]["actual_amount"] += float(res.actual_amount or 0)

    # --- write rows ---
    current_row = start_row
    header_row = current_row  # Track header row for grouping

    # sub-header row for purchased items
    sub_headers = list(PURCHASED_COLUMN_MAP.values())
    for col_idx, header_name in enumerate(sub_headers, start=1):
        cell = ws.cell(row=current_row, column=col_idx)
        cell.value = header_name
        cell.font = font
        cell.alignment = alignment
        cell.border = border
        cell.fill = blue_bg
    ws.row_dimensions[current_row].height = 20
    current_row += 1

    if not grouped:
        # no data: write placeholder
        cell = ws.cell(row=current_row, column=1)
        cell.value = "ไม่มีข้อมูลการจัดซื้อ"
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.merge_cells(
            start_row=current_row,
            start_column=1,
            end_row=current_row,
            end_column=len(sub_headers),
        )
        current_row += 1
        return current_row

    for req_id, data in grouped.items():
        req = data["requisition"]
        items = req.items if req.items else []
        num_items = max(len(items), 1)

        first_row = current_row
        last_row = current_row + num_items - 1

        # --- Column A: วันที่เอกสาร (order_confirmed timestamp) ---
        ts = data["order_confirmed_ts"]
        ts_str = ts.strftime("%d/%m/%Y") if ts else "-"
        ws.cell(row=first_row, column=1).value = ts_str
        if num_items > 1:
            ws.merge_cells(
                start_row=first_row, start_column=1, end_row=last_row, end_column=1
            )
        ws.cell(row=first_row, column=1).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # --- Column B: requisition_code ---
        ws.cell(row=first_row, column=2).value = req.requisition_code or "-"
        if num_items > 1:
            ws.merge_cells(
                start_row=first_row, start_column=2, end_row=last_row, end_column=2
            )
        ws.cell(row=first_row, column=2).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # --- Column D: total amount (merged for all items in this requisition) ---
        ws.cell(row=first_row, column=4).value = data["actual_amount"]
        ws.cell(row=first_row, column=4).number_format = "#,##0.00"
        ws.cell(row=first_row, column=4).alignment = Alignment(
            horizontal="right", vertical="center"
        )
        if num_items > 1:
            ws.merge_cells(
                start_row=first_row, start_column=4, end_row=last_row, end_column=4
            )

        # --- Column E: quotation_winner (merged) ---
        ws.cell(row=first_row, column=5).value = data["quotation_winner"] or "-"
        if num_items > 1:
            ws.merge_cells(
                start_row=first_row, start_column=5, end_row=last_row, end_column=5
            )
        ws.cell(row=first_row, column=5).alignment = Alignment(
            horizontal="center", vertical="center", wrap_text=True
        )

        # --- Column C: item names (one row per item) ---
        for idx, item in enumerate(items):
            row_idx = first_row + idx
            # item name
            ws.cell(row=row_idx, column=3).value = item.product_name or "-"
            ws.cell(row=row_idx, column=3).alignment = Alignment(
                horizontal="left", vertical="center", wrap_text=True
            )

        # apply thin border to all cells in the group
        thin = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for r in range(first_row, last_row + 1):
            for c in range(1, len(sub_headers) + 1):
                ws.cell(row=r, column=c).border = thin

        current_row = last_row + 1

    return current_row  # return next available row for the caller


def process_mas_export(current_user, start_date=None, end_date=None):
    # Convert strings to date objects if they are passed as strings
    if isinstance(start_date, str):
        try:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
        except ValueError:
            start_date = datetime.date.today().replace(month=1, day=1)

    if isinstance(end_date, str):
        try:
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        except ValueError:
            end_date = datetime.date.today().replace(month=12, day=31)

    # Fallback if None
    if not start_date:
        start_date = datetime.date.today().replace(month=1, day=1)
    if not end_date:
        end_date = datetime.date.today().replace(month=12, day=31)

    # 1. Query ข้อมูล
    mas_qs = models.MAS.objects(status="active")

    # 2. สร้าง Workbook และ Worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "ข้อมูลบุคลากร"

    # 3. กำหนด Header (ชื่อคอลัมน์)
    headers = list(EXPORT_MAS_COLUMN_MAP.values())
    year_range_text = format_year_range(start_date, end_date)
    header_text = (
        f"รายงานสรุปการใช้เงินทั้งหมด\n"
        f"{year_range_text}\n"
        "สำนักนวัตกรรมดิจิทัลและระบบอัจฉริยะ มหาวิทยาลัยสงขลานครินทร์\n"
        "ระยะเวลาตั้งแต่วันที่ {} ถึงวันที่ {}".format(
            format_date_th(start_date), format_date_th(end_date)
        )
    )
    ws.append(headers)
    merge_cells = f"A1:{get_column_letter(len(headers))}1"
    style_openpyxl_header(ws, headers, header_text, merge_cells)

    # 4. วนลูปใส่ข้อมูล
    for m in mas_qs:
        # write mas_code into column A
        print(f"Processing MAS: {m.id}")
        ws.append(
            [
                m.mas_code,
                m.description,
                float(m.amount or 0),
                float(m.amount - m.remaining_amount or 0),
                float(m.remaining_amount or 0),
            ]
        )
        row_idx = ws.max_row

        thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        for col in range(1, 6):
            cell = ws.cell(row=row_idx, column=col)
            cell.fill = light_blue_bg
            cell.border = thin_border

        # write purchased items for this MAS starting from next row
        next_row = ws.max_row + 1
        purchased_start_row = next_row  # Track where purchased items start
        next_row = get_purchased_items_in_mas(ws, m.id, next_row, start_date, end_date)
        purchased_end_row = next_row - 1  # Track where purchased items end

        # Group the purchased items rows under this MAS and ensure borders
        if purchased_end_row >= purchased_start_row:
            for r in range(purchased_start_row, purchased_end_row + 1):
                current_level = ws.row_dimensions[r].outline_level or 0
                ws.row_dimensions[r].outline_level = max(current_level, 1)
                ws.row_dimensions[r].hidden = True  # Collapse by default
                # Apply border to all cells in purchased items rows
                for col in range(1, 6):
                    cell = ws.cell(row=r, column=col)
                    if not cell.border or cell.border.left.style != "thin":
                        cell.border = thin_border

        # blank separator row between MAS blocks
        ws.append([""])

    # set format ทุก column เป็น text
    start_row = 2
    end_row = 10000
    for col_idx in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col_idx)
        # set ความกว้าง column
        ws.column_dimensions[col_letter].width = 30
        for row in range(start_row, end_row + 1):
            ws[f"{col_letter}{row}"].number_format = "@"

    # set format
    currency_format = "#,##0.00"
    column_formats = {
        "จำนวนเงินที่ขอจัดตั้ง": currency_format,
        "จำนวนเงินที่ใช้ไปแล้ว": currency_format,
        "จำนวนเงินคงเหลือ": currency_format,
    }

    for col_name, fmt in column_formats.items():
        col_idx = headers.index(col_name) + 1
        col_letter = get_column_letter(col_idx)
        ws.column_dimensions[col_letter].width = 20
        for row in range(start_row, end_row + 1):
            ws[f"{col_letter}{row}"].number_format = fmt
    if not wb.close:
        wb.close()

    # 5. Save ลง Memory Stream
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)  # เลื่อน pointer กลับไปที่จุดเริ่มต้นไฟล์
    export_mas_file = ExportFile.objects(created_by=current_user).first()

    # บันทึกไฟล์
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"ข้อมูล_MAS_{timestamp}.xlsx"

    if not export_mas_file:
        export_mas_file = ExportFile(
            created_date=datetime.datetime.now(),
            created_by=current_user,
        )
    export_mas_file.updated_date = datetime.datetime.now()
    export_mas_file.updater = current_user
    if not export_mas_file.file:
        export_mas_file.file.put(
            output,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        export_mas_file.file.replace(
            output,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    export_mas_file.file_name = filename
    export_mas_file.type_ = "mas_export"
    export_mas_file.status = "completed"
    export_mas_file.save()
    return True


# ส่งออก timeline items
def _write_timeline_item_row(ws, timeline_item, row_num, timeline=None):
    insurance_duration = timeline_item.get_format_insurance_duration()

    # Prepare row data - must match TIMELINE_ITEMS_COLUMN_MAP keys
    row_data = {
        "requisition_code": (
            timeline_item.requisition.requisition_code
            if timeline_item.requisition
            else "-"
        ),
        "running_number": timeline_item.running_number or "-",
        "requisition_item": timeline_item.requisition_item or "-",
        "seller": timeline_item.seller or "-",
        "serial_number": timeline_item.serial_number or "-",
        "requisition_item_code": timeline_item.requisition_item_code or "-",
        "insurance_start_date": timeline_item.insurance_start_date or "-",
        "insurance_end_date": timeline_item.insurance_end_date or "-",
        "insurance_duration": insurance_duration,
        "location": timeline_item.location or "-",
        "responder_user": (
            timeline_item.responder_user.get_name()
            if timeline_item.responder_user
            else "-"
        ),
    }

    # Write data to cells
    for col_idx, key in enumerate(TIMELINE_ITEMS_COLUMN_KEYS, start=1):
        cell = ws.cell(row=row_num, column=col_idx)
        cell.value = row_data.get(key, "-")
        cell.border = border
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)

        # Center alignment for specific columns
        if key in ["running_number"] or key == "requisition_code":
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )

        # Right alignment for numbers
        if key == "serial_number" or key == "requisition_item_code":
            cell.alignment = Alignment(
                horizontal="right", vertical="center", wrap_text=True
            )

    ws.row_dimensions[row_num].height = 25
    return row_num + 1


def requisition_timeline_items_export(current_user, start_date=None, end_date=None):
    # Parse dates
    if isinstance(start_date, str):
        try:
            start_date = datetime.datetime.strptime(start_date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            start_date = None

    if isinstance(end_date, str):
        try:
            end_date = datetime.datetime.strptime(end_date, "%Y-%m-%d").date()
        except (ValueError, TypeError):
            end_date = None

    # Set default dates if None
    if not start_date:
        start_date = datetime.date.today().replace(month=1, day=1)
    if not end_date:
        end_date = datetime.date.today().replace(month=12, day=31)

    # Query completed timelines first (those with progress_status = "completed")
    completed_timelines = models.RequisitionTimeline.objects(
        progress__progress_status="completed"
    )

    # Get their IDs for filtering timeline items
    completed_timeline_ids = [t.id for t in completed_timelines]

    # Query timeline items from completed timelines
    query = models.RequisitionTimelineItem.objects(
        status="active", requisition_timeline__in=completed_timeline_ids
    )

    # Filter by date range if provided
    if start_date and end_date:
        start_dt = datetime.datetime.combine(start_date, datetime.time.min)
        end_dt = datetime.datetime.combine(end_date, datetime.time.max)
        query = query.filter(created_date__gte=start_dt, created_date__lte=end_dt)

    # Order by requisition_code and running_number
    timeline_items = query.order_by("requisition__requisition_code", "running_number")
    total_count = timeline_items.count()
    print(
        f"Found {total_count} timeline items from {len(completed_timeline_ids)} completed timelines"
    )

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "รายงานรายการพัสดุ"

    # Write header using style_openpyxl_header
    headers = list(TIMELINE_ITEMS_COLUMN_MAP.values())
    year_range_text = format_year_range(start_date, end_date)
    header_text = (
        "รายงานรายการพัสดุและประกัน\n"
        f"{year_range_text}\n"
        "สำนักนวัตกรรมดิจิทัลและระบบอัจฉริยะ มหาวิทยาลัยสงขลานครินทร์\n"
        "ระยะเวลาตั้งแต่วันที่ {} ถึงวันที่ {}".format(
            format_date_th(start_date), format_date_th(end_date)
        )
    )
    merge_cells = f"A1:{get_column_letter(len(headers))}1"
    style_openpyxl_header(ws, headers, header_text, merge_cells)

    # Write data rows (starting from row 3 since headers are in row 2)
    current_row = 3
    for item in timeline_items:
        try:
            # Get associated timeline for progress check
            timeline = (
                models.RequisitionTimeline.objects(
                    id=item.requisition_timeline.id
                ).first()
                if item.requisition_timeline
                else None
            )

            current_row = _write_timeline_item_row(ws, item, current_row, timeline)

        except Exception as e:
            print(f"Error processing item {item.id}: {str(e)}")
            continue

    # Set column widths
    column_widths = {
        "A": 15,  # timeline_id
        "B": 8,  # running_number
        "C": 20,  # requisition_item
        "D": 30,  # seller
        "E": 12,  # serial_number
        "F": 12,  # requisition_item_code
        "G": 15,  # insurance_start_date
        "H": 15,  # insurance_end_date
        "I": 15,  # insurance_duration
        "J": 15,  # location
        "K": 30,  # responder_user
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    # Save to database
    export_file = ExportFile.objects(
        created_by=current_user, type_="requisition_items_export"
    ).first()

    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"รายการพัสดุ_{timestamp}.xlsx"

    if not export_file:
        export_file = ExportFile(
            created_date=datetime.datetime.now(),
            created_by=current_user,
            type_="requisition_items_export",
        )

    export_file.updated_date = datetime.datetime.now()
    export_file.status = "completed"

    if not export_file.file:
        export_file.file.put(
            output,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        export_file.file.replace(
            output,
            filename=filename,
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

    export_file.file_name = filename
    export_file.type_ = "requisition_items_export"
    export_file.save()

    print(f"Export completed: {filename}")
    print(f"Total items exported: {total_count}")
    return True
