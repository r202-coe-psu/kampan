import mongoengine as me
import pandas as pd
import datetime
import re
import io
from .. import models
from decimal import Decimal
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

MAS_COLUMN_MAP = {
    "year": "Year\nปีงบประมาณ",
    "mas_code": "Item Code\nรหัสงบประมาณ",
    "amount": "Amount\nจำนวนเงินที่ขอจัดตั้ง",
    "description": "Description\nรายละเอียด",
    "direction": "Direction\nทิศทาง",
}
MAS_COLUMNS = list(MAS_COLUMN_MAP.keys())

MA_COLUMN_MAP = {
    "product_number": "เลขที่สินค้า/เลขที่เอกสาร",
    "name": "ชื่อรายการ",
    "asset_code": "รหัสครุภัณฑ์",
    "start_date": "วันที่เริ่มต้น",
    "end_date": "วันที่สิ้นสุด",
    "amount": "จำนวนเงิน(บาท)",
    "period": "จำนวนงวด",
    "category": "ประเภท",
    "responsible_by": "ผู้รับผิดชอบ",
    "company": "บริษัท",
}
MA_COLUMNS = list(MA_COLUMN_MAP.keys())


def _normalize_columns(df: pd.DataFrame, column_map: dict) -> pd.DataFrame:
    reverse = {v: k for k, v in column_map.items()}
    return df.rename(columns=reverse)


def generate_mas_template() -> io.BytesIO:

    example_rows = [
        {
            "Year\nปีงบประมาณ": 2569,
            "Item Code\nรหัสงบประมาณ": "MAS-2024-001",
            "Description\nรายละเอียด": "งบดำเนินการ ประจำปี 2567",
            "Amount\nจำนวนเงินที่ขอจัดตั้ง": 500000.00,
            "Direction\nทิศทาง": "เข้า",
        },
        {
            "Year\nปีงบประมาณ": 2569,
            "Item Code\nรหัสงบประมาณ": "MAS-2024-002",
            "Description\nรายละเอียด": "งบลงทุน ครุภัณฑ์สำนักงาน",
            "Amount\nจำนวนเงินที่ขอจัดตั้ง": 1200000.00,
            "Direction\nทิศทาง": "ออก",
        },
    ]
    headers = list(MAS_COLUMN_MAP.values())
    df = pd.DataFrame(example_rows, columns=headers)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="MAS")
        wb = writer.book
        ws = writer.sheets["MAS"]

        header_fill = PatternFill("solid", fgColor="4472C4")
        example_fill = PatternFill("solid", fgColor="D9E1F2")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        example_font = Font(color="595959", italic=True, size=10)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = example_fill
                cell.font = example_font
                cell.alignment = Alignment(vertical="center")
                cell.border = border

        for col_idx, col in enumerate(headers, start=1):
            col_values = [col] + [
                str(example_rows[r].get(col, "")) for r in range(len(example_rows))
            ]
            max_len = max(len(v) for v in col_values) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len

        ws.row_dimensions[1].height = 22

    buf.seek(0)
    return buf


def generate_ma_template() -> io.BytesIO:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    example_rows = [
        {
            "เลขที่สินค้า/เลขที่เอกสาร": "DOC-2024-001",
            "ชื่อรายการ": "เครื่องคอมพิวเตอร์โน้ตบุ๊ก",
            "รหัสครุภัณฑ์": "IT-NB-0001",
            "วันที่เริ่มต้น": "01/01/2024",
            "วันที่สิ้นสุด": "31/12/2024",
            "จำนวนเงิน(บาท)": 35000.00,
            "จำนวนงวด": 12,
            "ประเภท": "ครุภัณฑ์",
            "ผู้รับผิดชอบ": "สมชาย ใจดี",
            "บริษัท": "บริษัท เทคโนโลยี จำกัด",
        },
        {
            "เลขที่สินค้า/เลขที่เอกสาร": "DOC-2024-002",
            "ชื่อรายการ": "ซอฟต์แวร์บัญชี",
            "รหัสครุภัณฑ์": "SW-ACC-0001",
            "วันที่เริ่มต้น": "01/03/2024",
            "วันที่สิ้นสุด": "28/02/2025",
            "จำนวนเงิน(บาท)": 120000.00,
            "จำนวนงวด": 1,
            "ประเภท": "ซอฟต์แวร์",
            "ผู้รับผิดชอบ": "สมหญิง รักสงบ",
            "บริษัท": "บริษัท ซอฟต์แวร์ไทย จำกัด",
        },
    ]
    headers = list(MA_COLUMN_MAP.values())
    df = pd.DataFrame(example_rows, columns=headers)

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="MA")
        wb = writer.book
        ws = writer.sheets["MA"]

        header_fill = PatternFill("solid", fgColor="375623")
        example_fill = PatternFill("solid", fgColor="E2EFDA")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        example_font = Font(color="595959", italic=True, size=10)
        center = Alignment(horizontal="center", vertical="center")
        thin = Side(style="thin", color="B0B0B0")
        border = Border(left=thin, right=thin, top=thin, bottom=thin)

        for col_idx, col in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = border

        for row_idx in range(2, len(df) + 2):
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = example_fill
                cell.font = example_font
                cell.alignment = Alignment(vertical="center")
                cell.border = border

        for col_idx, col in enumerate(headers, start=1):
            col_values = [col] + [
                str(example_rows[r].get(col, "")) for r in range(len(example_rows))
            ]
            max_len = max(len(v) for v in col_values) + 4
            ws.column_dimensions[get_column_letter(col_idx)].width = max_len

        ws.row_dimensions[1].height = 22

        note_row = len(df) + 3
        ws.cell(
            row=note_row,
            column=1,
            value="หมายเหตุ: ประเภท (ประเภท) ให้ระบุเป็น: ครุภัณฑ์, ซอฟต์แวร์, จ้างเหมาบริการ, วัสดุ",
        )
        ws.cell(row=note_row, column=1).font = Font(color="FF0000", italic=True, size=9)
        ws.cell(
            row=note_row + 1,
            column=1,
            value="หมายเหตุ: วันที่ให้ระบุในรูปแบบ DD/MM/YYYY เช่น 01/01/2024",
        )
        ws.cell(row=note_row + 1, column=1).font = Font(
            color="FF0000", italic=True, size=9
        )

    buf.seek(0)
    return buf


def to_decimal(val):
    try:
        return Decimal(str(val))
    except Exception:
        return Decimal("0.00")


def safe_split(value, sep=","):
    if isinstance(value, str):
        return [v.strip() for v in value.split(sep)]
    elif value is None or (isinstance(value, float) and pd.isna(value)):
        return []
    else:
        return [v.strip() for v in str(value).split(sep)]


def safe_str(value):
    if isinstance(value, str):
        return value.strip()
    elif value is None or (isinstance(value, float) and pd.isna(value)):
        return ""
    else:
        return str(value).strip()


def is_missing_required(value):
    return (
        value is None
        or (isinstance(value, float) and pd.isna(value))
        or str(value).strip() == ""
    )


def validate_mas_file(document) -> list:

    errors = []
    try:
        document.file.seek(0)
        df = pd.read_excel(document.file)
        df.columns = df.columns.str.strip()
        df = _normalize_columns(df, MAS_COLUMN_MAP)
    except Exception as e:
        return [f"ไม่สามารถอ่านไฟล์ได้: {e}"]

    missing_cols = set(MAS_COLUMNS) - set(df.columns)
    if missing_cols:
        thai_missing = [MAS_COLUMN_MAP.get(c, c) for c in missing_cols]
        return [f"ไม่พบคอลัมน์ที่จำเป็น: {', '.join(thai_missing)}"]

    for idx, row in df.iterrows():
        row_num = idx + 2
        missing = [
            MAS_COLUMN_MAP.get(k, k)
            for k in MAS_COLUMNS
            if is_missing_required(row.get(k))
        ]
        if missing:
            errors.append(f"แถวที่ {row_num}: ขาดข้อมูล {', '.join(missing)}")
            continue

        mas_code = safe_str(row.get("mas_code"))
        if not mas_code:
            errors.append(f"แถวที่ {row_num}: รหัสแหล่งเงินว่างเปล่า")
            continue

        if models.MAS.objects(mas_code=mas_code, status="active").first():
            errors.append(f"แถวที่ {row_num}: รหัสแหล่งเงิน '{mas_code}' มีอยู่ในระบบแล้ว")

        year = safe_str(row.get("year"))
        if not year:
            errors.append(f"แถวที่ {row_num}: ปีงบประมาณว่างเปล่า")
            continue

        description = safe_str(row.get("description"))
        if len(description) < 3:
            errors.append(f"แถวที่ {row_num}: รายละเอียดสั้นเกินไป ('{description}')")

        direction = safe_str(row.get("direction"))
        if not direction:
            errors.append(f"แถวที่ {row_num}: ทิศทางว่างเปล่า")
            continue

        for col, label in [
            ("amount", MAS_COLUMN_MAP["amount"]),
        ]:
            val = row.get(col)
            if not is_missing_required(val):
                try:
                    if float(val) < 0:
                        errors.append(f"แถวที่ {row_num}: {label} ต้องไม่ติดลบ")
                except (ValueError, TypeError):
                    errors.append(f"แถวที่ {row_num}: {label} ต้องเป็นตัวเลข")

    return errors


def validate_ma_file(document) -> list:

    errors = []
    try:
        document.file.seek(0)
        file_bytes = document.file.read()
        df = pd.read_excel(io.BytesIO(file_bytes))
        df.columns = df.columns.str.strip()
        df = _normalize_columns(df, MA_COLUMN_MAP)
    except Exception as e:
        return [f"ไม่สามารถอ่านไฟล์ได้: {e}"]

    missing_cols = set(MA_COLUMNS) - set(df.columns)
    if missing_cols:
        thai_missing = [MA_COLUMN_MAP.get(c, c) for c in missing_cols]
        return [f"ไม่พบคอลัมน์ที่จำเป็น: {', '.join(thai_missing)}"]

    for idx, row in df.iterrows():
        row_num = idx + 2
        missing = [
            MA_COLUMN_MAP.get(k, k)
            for k in MA_COLUMNS
            if is_missing_required(row.get(k))
        ]
        if missing:
            errors.append(f"แถวที่ {row_num}: ขาดข้อมูล {', '.join(missing)}")
            continue

        if pd.notna(row.get("start_date")) and pd.notna(row.get("end_date")):
            try:
                sd = parse_date(row.get("start_date"))
                ed = parse_date(row.get("end_date"))
                if sd and ed and ed < sd:
                    errors.append(f"แถวที่ {row_num}: วันที่สิ้นสุดต้องมากกว่าวันที่เริ่มต้น")
            except Exception:
                pass

        if pd.notna(row.get("amount")):
            try:
                if float(row.get("amount")) < 0:
                    errors.append(
                        f"แถวที่ {row_num}: {MA_COLUMN_MAP['amount']} ต้องไม่ติดลบ"
                    )
            except (ValueError, TypeError):
                errors.append(f"แถวที่ {row_num}: {MA_COLUMN_MAP['amount']} ต้องเป็นตัวเลข")

        if pd.notna(row.get("period")):
            try:
                if int(float(row.get("period"))) < 1:
                    errors.append(
                        f"แถวที่ {row_num}: {MA_COLUMN_MAP['period']} ต้องมากกว่า 0"
                    )
            except (ValueError, TypeError):
                errors.append(f"แถวที่ {row_num}: {MA_COLUMN_MAP['period']} ต้องเป็นตัวเลข")

    return errors


def save_mas_db(document, mas, user_id):
    print("=====> Start processing MAS document:", document.id)
    """
    อ่านข้อมูล MAS จากไฟล์ Excel แล้วสร้าง MAS แต่ละตัวลงฐานข้อมูล

    Expected columns: mas_code, description, amount, direction
    """
    errors = []
    processed_user = models.User.objects(id=user_id).first()
    document = models.Document.objects(id=document.id).first()
    if not document:
        print(f"Document not found.")
        return False, []

    # Clear previous errors
    document.error_messages = []

    try:
        document.file.seek(0)
        df = pd.read_excel(document.file)
        df.columns = df.columns.str.strip()
        df = _normalize_columns(df, MAS_COLUMN_MAP)
    except Exception as e:
        document.status = "failed"
        document.error_messages = [f"ไม่สามารถอ่านไฟล์ได้: {e}"]
        document.save()
        print(f"Error reading Excel in save_mas_db: {e}")
        return False, document.error_messages

    print(f"\n==== Loaded Excel File: {len(df)} rows ====\n")
    if not processed_user:
        document.status = "failed"
        document.error_messages = ["ไม่พบข้อมูลผู้ใช้"]
        document.save()
        print("Cannot find user.")
        return False, document.error_messages

    required_cols = {
        "year",
        "mas_code",
        "description",
        "amount",
        "direction",
    }
    missing_cols = required_cols - set(col.lower() for col in df.columns)
    if missing_cols:
        thai_missing = [MAS_COLUMN_MAP.get(c, c) for c in missing_cols]
        msg = f"ไม่พบคอลัมน์ที่จำเป็น: {', '.join(thai_missing)}"
        document.status = "failed"
        document.error_messages = [msg]
        document.save()
        print(msg)
        return False, [msg]

    created_count = 0

    for i, row in df.iterrows():
        row_num = i + 2
        required_fields = {
            "year": row.get("year"),
            "mas_code": row.get("mas_code"),
            "description": row.get("description"),
            "amount": row.get("amount"),
            "direction": row.get("direction"),
        }
        missing = [
            MAS_COLUMN_MAP.get(k, k)
            for k, v in required_fields.items()
            if is_missing_required(v)
        ]
        if missing:
            msg = f"แถวที่ {row_num}: ขาดข้อมูล {', '.join(missing)}"
            errors.append(msg)
            print(msg)
            continue

        year = to_decimal(row.get("year"))
        if year < 2500 or year > 2700:
            msg = f"แถวที่ {row_num}: ปีงบประมาณต้องอยู่ระหว่าง พ.ศ. 2500 ถึง พ.ศ. 2700 (พบ {year})"
            errors.append(msg)
            print(msg)
            continue

        description = safe_str(row.get("description"))
        if len(description) < 3:
            msg = f"แถวที่ {row_num}: รายละเอียดสั้นเกินไป ('{description}')"
            errors.append(msg)
            print(msg)
            continue

        mas_code = safe_str(row.get("mas_code"))
        if not mas_code:
            msg = f"แถวที่ {row_num}: รหัสแหล่งเงินว่างเปล่า"
            errors.append(msg)
            print(msg)
            continue

        direction = safe_str(row.get("direction"))
        if not direction:
            msg = f"แถวที่ {row_num}: ทิศทางว่างเปล่า"
            errors.append(msg)
            print(msg)
            continue

        existing = models.MAS.objects(mas_code=mas_code, status="active").first()
        if existing:
            msg = f"แถวที่ {row_num}: รหัสแหล่งเงิน '{mas_code}' มีอยู่ในระบบแล้ว"
            errors.append(msg)
            print(msg)
            continue

        amount = to_decimal(row.get("amount"))
        remaining_amount = to_decimal(row.get("remaining_amount"))
        reservable_amount = to_decimal(row.get("reservable_amount"))

        mas_obj = models.MAS(
            year=year,
            mas_code=mas_code,
            description=description,
            amount=amount,
            remaining_amount=amount,
            reservable_amount=amount,
            direction=direction,
            status="active",
            created_by=processed_user,
            last_updated_by=processed_user,
            created_date=datetime.datetime.now(),
            updated_date=datetime.datetime.now(),
        )
        mas_obj.save()
        created_count += 1
        print(f"Created MAS #{i + 1}: {mas_obj.description} ({mas_obj.mas_code})")

    document.updated_date = datetime.datetime.now()
    document.error_messages = errors
    if created_count == 0:
        document.status = "failed"
    elif created_count != len(df):
        document.status = "incomplete"
    else:
        document.status = "completed"
    document.save()
    print(f"Document status: {document.status}")
    print(f"Total MASs created: {created_count}/{len(df)}")

    return True, errors


def parse_date(val):
    if isinstance(val, str):
        val = val.strip()
        th_to_en_month = {
            "มกราคม": "January",
            "กุมภาพันธ์": "February",
            "มีนาคม": "March",
            "เมษายน": "April",
            "พฤษภาคม": "May",
            "มิถุนายน": "June",
            "กรกฎาคม": "July",
            "สิงหาคม": "August",
            "กันยายน": "September",
            "ตุลาคม": "October",
            "พฤศจิกายน": "November",
            "ธันวาคม": "December",
        }
        m = re.match(r"(\d{1,2})\s+([ก-๙]+)\s+(\d{4})", val)
        if m:
            day = m.group(1)
            th_month = m.group(2)
            year = int(m.group(3)) - 543
            en_month = th_to_en_month.get(th_month)
            if en_month:
                date_str = f"{day} {en_month} {year}"
                try:
                    from datetime import datetime

                    dt = datetime.strptime(date_str, "%d %B %Y")
                    return dt
                except Exception:
                    pass
        try:
            return pd.to_datetime(val, errors="coerce")
        except Exception:
            return None
    else:
        return pd.to_datetime(val, errors="coerce")


def save_ma_db(document, ma, user_id):
    print("=====> Start processing MA document:", document.id)
    """
    อ่านข้อมูล MA จากไฟล์ Excel แล้วสร้าง MA แต่ละตัวลงฐานข้อมูล
    """
    errors = []
    processed_user = models.User.objects(id=user_id).first()
    document = models.Document.objects(id=document.id).first()
    created_count = 0
    if not document:
        print(f"Document not found.")
        return False

    # Clear previous errors
    document.error_messages = []

    try:
        document.file.seek(0)
        file_bytes = document.file.read()
        df = pd.read_excel(io.BytesIO(file_bytes))
        df.columns = df.columns.str.strip()
        df = _normalize_columns(df, MA_COLUMN_MAP)
    except Exception as e:
        document.status = "failed"
        document.error_messages = [f"ไม่สามารถอ่านไฟล์ได้: {e}"]
        document.save()
        print(f"Error reading Excel in save_ma_db: {e}")
        return False

    print(f"\n==== Loaded Excel File: {len(df)} rows ====\n")
    if not processed_user:
        document.status = "failed"
        document.error_messages = ["ไม่พบข้อมูลผู้ใช้"]
        document.save()
        print("Cannot find user.")
        return False

    column_map = {
        "product_number": "product_number",
        "name": "name",
        "asset_code": "asset_code",
        "start_date": "start_date",
        "end_date": "end_date",
        "amount": "amount",
        "period": "period",
        "category": "category",
        "responsible_by": "responsible_by",
        "company": "company",
    }
    category_map = {
        "ซอฟต์แวร์": "software",
        "ครุภัณฑ์": "product",
        "จ้างเหมาบริการ": "service",
        "วัสดุ": "material",
    }

    required_columns = list(column_map.keys())

    for idx, row in df.iterrows():
        row_num = idx + 2
        missing = False
        for col in required_columns:
            if pd.isnull(row.get(col)) or str(row.get(col)).strip() == "":
                msg = f"แถวที่ {row_num}: ขาดข้อมูล '{MA_COLUMN_MAP.get(col, col)}'"
                print(msg)
                errors.append(msg)
                missing = True
                break
        if missing:
            continue

        data = {}
        try:
            for excel_col, model_field in column_map.items():
                value = row.get(excel_col)
                try:
                    if model_field in ["start_date", "end_date"]:
                        value = parse_date(value)
                    if model_field == "amount":
                        value = (
                            to_decimal(value) if pd.notnull(value) else Decimal("0.00")
                        )
                    if model_field == "period":
                        value = int(value) if pd.notnull(value) else 1
                    if model_field == "asset_code":
                        value = str(value) if pd.notnull(value) else ""
                    if model_field == "category":
                        value = category_map.get(str(value).strip(), "other")
                    if model_field == "responsible_by":
                        responsible_by_list = []
                        names = str(value).strip().splitlines()
                        for name in names:
                            name = name.strip()
                            if not name:
                                continue
                            fullname = name.split(" ", 1)
                            if len(fullname) == 2:
                                first_name, last_name = fullname
                                user_role = models.OrganizationUserRole.objects(
                                    first_name=first_name.strip(),
                                    last_name=last_name.strip(),
                                ).first()
                                if user_role:
                                    responsible_by_list.append(user_role)
                        data[model_field] = responsible_by_list
                        continue
                    data[model_field] = value
                except Exception as field_e:
                    msg = f"แถวที่ {row_num}: คอลัมน์ '{MA_COLUMN_MAP.get(excel_col, excel_col)}' ผิดพลาด: {field_e}"
                    print(msg)
                    errors.append(msg)

            duplicate_query = {
                "name": data.get("name"),
                "asset_code": data.get("asset_code"),
                "start_date": data.get("start_date"),
                "end_date": data.get("end_date"),
                "amount": data.get("amount"),
                "period": data.get("period"),
                "company": data.get("company"),
                "category": data.get("category"),
                "responsible_by": data.get("responsible_by"),
            }
            if models.Procurement.objects(**duplicate_query).first():
                msg = f"แถวที่ {row_num}: ข้อมูลซ้ำกับที่มีในระบบแล้ว"
                print(msg)
                errors.append(msg)
                continue

            procurement = models.Procurement(
                **data,
                created_by=processed_user,
                last_updated_by=processed_user,
                status="active",
            )
            procurement.save()
            created_count += 1
            print(f"Saved procurement: {procurement.product_number}")
        except Exception as e:
            msg = f"แถวที่ {row_num}: เกิดข้อผิดพลาด: {e}"
            print(msg)
            errors.append(msg)
            continue

    document.updated_date = datetime.datetime.now()
    document.error_messages = errors
    if created_count == 0:
        document.status = "failed"
    elif created_count < len(df):
        document.status = "incomplete"
    else:
        document.status = "completed"
    document.save()
    print(f"Document status: {document.status}")
    print(f"Total MAs created: {created_count}/{len(df)}")

    return True
